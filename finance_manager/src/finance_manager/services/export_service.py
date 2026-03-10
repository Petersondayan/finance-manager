"""Export service — generate PDF and XLSX reports."""

import calendar
from pathlib import Path
from .report_service import MonthlyReport, AnnualReport


class ExportService:
    """Export financial reports to PDF and XLSX."""

    # ----- XLSX exports -----

    def export_monthly_xlsx(self, report: MonthlyReport, path: str) -> None:
        """Write monthly report to an Excel file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = f"Report {report.period}"

        ws["A1"] = f"Monthly Report — {report.period}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        ws.append([])
        ws.append(["Income", report.income])
        ws.append(["Expenses", report.expenses])
        ws.append(["Net Change", report.net_change])
        ws.append(["Ending Balance", report.ending_balance])

        ws.append([])
        ws.append(["Category", "Amount"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for cat in report.top_categories:
            ws.append([cat["name"], cat["amount"]])

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18

        wb.save(path)

    def export_annual_xlsx(self, report: AnnualReport, path: str) -> None:
        """Write annual report to an Excel file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = f"Annual {report.year}"

        ws["A1"] = f"Annual Report — {report.year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        ws.append([])
        ws.append(["Total Income", report.total_income])
        ws.append(["Total Expenses", report.total_expenses])
        ws.append(["Net Savings", report.net_savings])
        ws.append(["Savings Rate", f"{report.savings_rate:.1f}%"])

        ws.append([])
        ws.append(["Month", "Income", "Expenses", "Net"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for row in report.monthly_breakdown:
            ws.append([
                calendar.month_name[row["month"]],
                row["income"],
                row["expenses"],
                row["net"],
            ])

        ws.append([])
        ws.append(["Category", "Amount"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for cat in report.category_breakdown:
            ws.append([cat["name"], cat["amount"]])

        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 22

        wb.save(path)

    # ----- PDF exports -----

    def export_monthly_pdf(self, report: MonthlyReport, path: str) -> None:
        """Write monthly report to a PDF file."""
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        doc = SimpleDocTemplate(path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"Monthly Report — {report.period}", styles["Title"]))
        story.append(Spacer(1, 12))

        summary_data = [
            ["Summary", ""],
            ["Income", f"${report.income:,.2f}"],
            ["Expenses", f"${report.expenses:,.2f}"],
            ["Net Change", f"${report.net_change:,.2f}"],
            ["Ending Balance", f"${report.ending_balance:,.2f}"],
        ]
        t = Table(summary_data, colWidths=[200, 150])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        story.append(t)
        story.append(Spacer(1, 20))

        if report.top_categories:
            story.append(Paragraph("Top Categories", styles["Heading2"]))
            cat_data = [["Category", "Amount"]] + [
                [c["name"], f"${c['amount']:,.2f}"] for c in report.top_categories
            ]
            ct = Table(cat_data, colWidths=[250, 150])
            ct.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(ct)

        doc.build(story)

    def export_annual_pdf(self, report: AnnualReport, path: str) -> None:
        """Write annual report to a PDF file."""
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        doc = SimpleDocTemplate(path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"Annual Report — {report.year}", styles["Title"]))
        story.append(Spacer(1, 12))

        summary_data = [
            ["Total Income", f"${report.total_income:,.2f}"],
            ["Total Expenses", f"${report.total_expenses:,.2f}"],
            ["Net Savings", f"${report.net_savings:,.2f}"],
            ["Savings Rate", f"{report.savings_rate:.1f}%"],
        ]
        t = Table(summary_data, colWidths=[200, 150])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        story.append(t)
        story.append(Spacer(1, 20))

        story.append(Paragraph("Monthly Breakdown", styles["Heading2"]))
        monthly_data = [["Month", "Income", "Expenses", "Net"]] + [
            [calendar.month_abbr[row["month"]],
             f"${row['income']:,.2f}", f"${row['expenses']:,.2f}", f"${row['net']:,.2f}"]
            for row in report.monthly_breakdown
        ]
        mt = Table(monthly_data, colWidths=[100, 120, 120, 120])
        mt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        story.append(mt)

        doc.build(story)
